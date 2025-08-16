import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from tkinter import Tk, filedialog, simpledialog, Label, Entry, Button
from datetime import datetime
import os

# -------------------- Constants --------------------
CANONICAL_REPORT_HEADERS = [
    "SAMPLE ID", "CHEMICAL", "CONCENTRATION", "pH LEVEL",
    "TEMPERATURE", "PRESSURE", "FLOW RATE", "STATUS", "COMMENT"
]

CSV_HEADER_ALIASES = {
    "sample_id|sampleid|id|sample": "SAMPLE ID",
    "chemical|compound|analyte|reagent": "CHEMICAL",
    "concentration_ppm|concentration|conc_ppm|conc|concentration_ppm_": "CONCENTRATION",
    "ph_level|ph": "pH LEVEL",
    "temperature_celsius|temperature_c|temp_c|temperature|temp": "TEMPERATURE",
    "pressure_kpa|pressure": "PRESSURE",
    "flow_rate_l_min|flowrate_l_min|flow_rate|flowrate|flow": "FLOW RATE"
}

RANGE_COL_MAP = {
    "chemical": "Chemical",
    "concentration_ppm_min": "Concentration_ppm_Min",
    "concentration_ppm_max": "Concentration_ppm_Max",
    "ph_level_min": "pH_Level_Min",
    "ph_level_max": "pH_Level_Max",
    "temperature_c_min": "Temperature_C_Min",
    "temperature_c_max": "Temperature_C_Max",
    "pressure_kpa_min": "Pressure_kPa_Min",
    "pressure_kpa_max": "Pressure_kPa_Max",
    "flow_rate_l_min_min": "Flow_Rate_L_min_Min",
    "flow_rate_l_min_max": "Flow_Rate_L_min_Max"
}

# -------------------- Utilities --------------------
def _norm(s: str) -> str:
    """Normalize column name for matching."""
    s = (s or "").strip().lower()
    replacements = {
        "°c|(c)|c°": "celsius", " c ": " celsius ",
        "kpa": "kpa", "l/min|l per min|l per minute": "l_min", "ph": "ph"
    }
    for old, new in replacements.items():
        for o in old.split("|"):
            s = s.replace(o, new)
    return "".join(c if c.isalnum() or c == "_" else "_" for c in s).strip("_")

# -------------------- File Handling --------------------
def get_user_info():
    info = {}

    def on_submit():
        info["FirstName"] = entry_first.get()
        info["MiddleName"] = entry_middle.get()
        info["LastName"] = entry_last.get()
        info["FirstIntl"] = entry_first.get()[:1]
        info["MidIntl"] = entry_middle.get()[:1]
        info["LastIntl"] = entry_last.get()[:1]
        info["CompanyName"] = entry_company.get()
        info["DateToday"] = datetime.today().strftime("%Y%m%d")

        root.quit()    # exit the Tk loop
        root.destroy() # close the window

    root = Tk()
    root.title("User Info")

    Label(root, text="First Name:").grid(row=0, column=0, sticky="e")
    entry_first = Entry(root)
    entry_first.grid(row=0, column=1)

    Label(root, text="Middle Name:").grid(row=1, column=0, sticky="e")
    entry_middle = Entry(root)
    entry_middle.grid(row=1, column=1)

    Label(root, text="Last Name:").grid(row=2, column=0, sticky="e")
    entry_last = Entry(root)
    entry_last.grid(row=2, column=1)

    Label(root, text="Company Name:").grid(row=3, column=0, sticky="e")
    entry_company = Entry(root)
    entry_company.grid(row=3, column=1)

    Button(root, text="Submit", command=on_submit).grid(row=4, column=0, columnspan=2, pady=10)

    root.mainloop()
    return info

def select_file(title="Select CSV File", filetypes=[("CSV Files", "*.csv")]):
    """Select file via dialog."""
    Tk().withdraw()
    return filedialog.askopenfilename(title=title, filetypes=filetypes)

def get_save_path(user_info):
    """Generate and verify save path."""
    initials = f"{user_info['FirstIntl']}{user_info['MidIntl']}{user_info['LastIntl']}"
    base_name = f"Chemical_Compliance_Report_{user_info['CompanyName']}_{user_info['DateToday']}_{initials}.xlsx"
    Tk().withdraw()
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=base_name)
    if save_path and os.path.exists(save_path):
        version = simpledialog.askstring("Version", "Enter version number:") or "2"
        base_name = f"Chemical_Compliance_Report_{user_info['CompanyName']}_{user_info['DateToday']}_{initials}_v{version}.xlsx"
        save_path = os.path.join(os.path.dirname(save_path), base_name)
    return save_path

# -------------------- Data Processing --------------------
def standardize_csv_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Standardize CSV headers and ensure required columns."""
    rename_map = {}
    used_targets = set()
    for col in df.columns:
        for aliases, target in CSV_HEADER_ALIASES.items():
            if _norm(col) in aliases.split("|") and target not in used_targets:
                rename_map[col] = target
                used_targets.add(target)
    
    df = df.rename(columns=rename_map)
    for col in CANONICAL_REPORT_HEADERS:
        if col not in df.columns:
            df[col] = pd.NA
    df = df[[c for c in CANONICAL_REPORT_HEADERS if c in df.columns] + 
            [c for c in df.columns if c not in CANONICAL_REPORT_HEADERS]]
    
    for col in ["CONCENTRATION", "pH LEVEL", "TEMPERATURE", "PRESSURE", "FLOW RATE"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

def load_ranges():
    """Load and standardize ranges from Excel."""
    ranges_path = r"C:\Users\mkb00\PROJECTS\PythonProjects\ComplianceMole\CompliantRanges.xlsx"
    rng = pd.read_excel(ranges_path)
    norm_to_original = {_norm(col): col for col in rng.columns}
    
    chem_col = next((norm_to_original[cand] for cand in ("chemical", "chemicals", "compound", "analyte") 
                if cand in norm_to_original), None)
    if not chem_col:
        raise KeyError("Could not find a 'Chemical' column in CompliantRanges.xlsx")

    
    rng = rng.rename(columns={orig: pretty for norm, pretty in RANGE_COL_MAP.items() 
                             if norm in norm_to_original for orig in [norm_to_original[norm]]})
    
    missing = [v for k, v in RANGE_COL_MAP.items() if k != "chemical" and v not in rng.columns]
    if missing:
        raise KeyError(f"Missing columns in CompliantRanges.xlsx: {', '.join(missing)}")
    
    return rng.set_index(chem_col)

def check_compliance(df: pd.DataFrame, ranges_df: pd.DataFrame) -> pd.DataFrame:
    """Check compliance against ranges."""
    req_cols = ["CHEMICAL", "CONCENTRATION", "pH LEVEL", "TEMPERATURE", "PRESSURE", "FLOW RATE"]
    if not all(c in df.columns for c in req_cols):
        df["STATUS"] = "UNKNOWN"
        df["COMMENT"] = "Required columns missing in source CSV."
        return df

    for idx, row in df.iterrows():
        chem = row["CHEMICAL"]
        if pd.isna(chem) or chem not in ranges_df.index:
            df.at[idx, "STATUS"] = "UNKNOWN CHEMICAL"
            df.at[idx, "COMMENT"] = "No compliance data found."
            continue

        limits = ranges_df.loc[chem]
        issues = []
        checks = [
            ("CONCENTRATION", "Concentration_ppm", row["CONCENTRATION"]),
            ("pH LEVEL", "pH_Level", row["pH LEVEL"]),
            ("TEMPERATURE", "Temperature_C", row["TEMPERATURE"]),
            ("PRESSURE", "Pressure_kPa", row["PRESSURE"]),
            ("FLOW RATE", "Flow_Rate_L_min", row["FLOW RATE"])
        ]
        
        for col, prefix, val in checks:
            try:
                if not (limits[f"{prefix}_Min"] <= val <= limits[f"{prefix}_Max"]):
                    issues.append(f"{col} not within acceptable range: {limits[f'{prefix}_Min']} - {limits[f'{prefix}_Max']}.")
            except (TypeError, ValueError):
                issues.append(f"{col} not within acceptable range: {limits[f'{prefix}_Min']} - {limits[f'{prefix}_Max']}.")
        
        df.at[idx, "STATUS"] = "NON-COMPLIANT" if issues else "COMPLIANT"
        df.at[idx, "COMMENT"] = " ".join(issues) if issues else "Within Acceptable Ranges"
    
    return df

# -------------------- Excel Formatting --------------------
#summary_ws = wb.create_sheet("Sample Data")
def format_excel(df, save_path, user_info):
    """Format the Excel output."""
    # Create workbook with Summary as the first sheet
    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    # Create Sample Data sheet second
    ws = wb.create_sheet("Sample Data")
    
    # Write DataFrame directly to Sample Data sheet using openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx).value = value
    
    # Save workbook to ensure sheets are created
    wb.save(save_path)
    
    # Debug: Verify sheet names after saving
    wb = openpyxl.load_workbook(save_path)
    print("Sheets in workbook:", wb.sheetnames)
    
    # Access sheets
    try:
        ws = wb["Sample Data"]
        summary_ws = wb["Summary"]
    except KeyError as e:
        raise KeyError(f"Sheet not found: {e}. Available sheets: {wb.sheetnames}")
    
    # Styles
    header_fill = PatternFill(start_color="5C6586", end_color="5C6586", fill_type="solid")
    subheader_fill = PatternFill(start_color="ADADAD", end_color="ADADAD", fill_type="solid")
    light_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin = Side(style="thin", color="000000")
    thick = Side(style="thick", color="000000")
    
    # Auto column width for Sample Data
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_length

    # Sample Data: Title row
    ws.insert_rows(1)
    ws.merge_cells("A1:I1")
    ws["A1"] = "SAMPLE DATA"
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(name="Aptos Display", bold=True, color="FFFFFF", size=12)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 20

    # Sample Data: Header row
    for col in range(1, 10):
        cell = ws.cell(2, col)
        cell.fill = light_fill
        cell.font = Font(name="Aptos Narrow", bold=True, size=10.5)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="double", color="000000"))
    ws.row_dimensions[2].height = 35

    # Sample Data: Data rows
    for row in range(3, ws.max_row + 1):
        for col in range(1, 10):
            cell = ws.cell(row, col)
            cell.font = Font(name="Aptos Narrow", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = white_fill
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws[f"I{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 15

    # Outline borders for Sample Data
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=9):
        for cell in row:
            borders = {k: cell.border.__dict__[k] for k in ("left", "right", "top", "bottom")}
            if cell.row == 1: borders["top"] = thick
            if cell.row == ws.max_row: borders["bottom"] = thick
            if cell.column == 1: borders["left"] = thick
            if cell.column == 9: borders["right"] = thick
            cell.border = Border(**borders)

    # Summary Sheet
    styles = {
        "header": (header_fill, Font(color="FFFFFF", bold=True), Alignment(horizontal="left", vertical="center")),
        "subheader": (subheader_fill, Font(color="000000", bold=True), Alignment(horizontal="center", vertical="center")),
        "light": (light_fill, Font(color="000000", bold=True), Alignment(horizontal="left", vertical="center")),
        "white": (white_fill, Font(color="000000"), Alignment(horizontal="right", vertical="center"))
    }

    # Fill empty cells in A1:Q(last_row+2) with white
    chemicals = sorted(df["CHEMICAL"].unique())
    last_row = 11 + len(chemicals) + 1 + 2 + len(chemicals) + 3  # Chemicals table + ranges table + spacing
    for row in range(1, last_row + 3):
        for col in range(1, 18):  # A to Q
            cell = summary_ws.cell(row, col)
            if not cell.value and not cell.fill.fgColor.rgb:
                cell.fill = white_fill

    # Summary Info
    summary_ws.merge_cells("B1:Q1")
    summary_ws["B1"] = "COMPLIANCE ANALYSIS REPORT"
    for fill, font, align in [styles["header"]]:
        summary_ws["B1"].fill, summary_ws["B1"].font, summary_ws["B1"].alignment = fill, font, align

    summary_info = {
        "B3": "Completed By:", "D3": "MK Barriault",
        "B4": "Date:", "D4": "8/15/2025",
        "B5": "Company:", "D5": "Ion Labs",
        "B6": "Total Samples:", "D6": "100",
        "B7": "Score:", "D7": ""
    }
    units_info = {
        "N3": "UNITS OF MEASURE", "N4": "Concentration =", "P4": "ppm",
        "N5": "Temperature =", "P5": "Celcius", "N6": "Pressure =", "P6": "kPa",
        "N7": "FlowRate =", "P7": "L/min"
    }
    
    # Apply merges for summary info and units
    merge_ranges = [
        "B3:C3", "B4:C4", "B5:C5", "B6:C6", "B7:C7",
        "D3:E3", "D4:E4", "D5:E5", "D6:E6", "D7:E7",
        "N3:Q3", "N4:O4", "N5:O5", "N6:O6", "N7:O7",
        "P4:Q4", "P5:Q5", "P6:Q6", "P7:Q7"
    ]
    for rng in merge_ranges:
        summary_ws.merge_cells(rng)
    
    for cell, value in {**summary_info, **units_info}.items():
        summary_ws[cell] = value
        style_key = "light" if cell.startswith(("B", "N")) else "white"
        if cell == "N3": style_key = "subheader"
        for fill, font, align in [styles[style_key]]:
            summary_ws[cell].fill, summary_ws[cell].font, summary_ws[cell].alignment = fill, font, align

    # Chemicals Table
    start_row = 11
    summary_ws.merge_cells("B9:Q9")
    summary_ws["B9"] = "ANALYSIS SUMMARY"
    for fill, font, align in [styles["header"]]:
        summary_ws["B9"].fill, summary_ws["B9"].font, summary_ws["B9"].alignment = fill, font, align

    categories = ["TOTAL SAMPLES", "PASS", "FAIL", "SCORE", "PRIORITY"]
    blocks = [(3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]
    for cat, (start, end) in zip(categories, blocks):
        summary_ws.merge_cells(start_row=10, start_column=start, end_row=10, end_column=end)
        cell = summary_ws.cell(10, start)
        cell.value, cell.fill, cell.font, cell.alignment = cat, *styles["subheader"]

    row = start_row
    for chem in chemicals:
        chem_df = df[df["CHEMICAL"] == chem]
        total, acceptable = len(chem_df), len(chem_df[chem_df["STATUS"] == "COMPLIANT"])
        percent = acceptable / total if total else 0
        values = [
            (3, 4, total), (5, 6, acceptable), (7, 8, total - acceptable),
            (9, 10, percent, "0.00%"), (11, 12, "HIGH" if percent < 0.45 else "LOW" if percent > 0.55 else "MEDIUM")
        ]
        
        summary_ws[f"B{row}"] = chem
        summary_ws[f"B{row}"].fill, summary_ws[f"B{row}"].font, summary_ws[f"B{row}"].alignment = styles["light"]
        
        for start, end, val, *fmt in values:
            summary_ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            cell = summary_ws.cell(row, start)
            cell.value, cell.fill, cell.alignment = val, white_fill, Alignment(horizontal="center", vertical="center")
            if fmt: cell.number_format = fmt[0]
        
        row += 1

    # Totals Row
    summary_ws[f"B{row}"] = "TOTAL:"
    for start, end, formula in [(3, 4, f"SUM(C{start_row}:C{row-1})"), (5, 6, f"SUM(E{start_row}:E{row-1})"), 
                               (7, 8, f"SUM(G{start_row}:G{row-1})"), (9, 10, None)]:
        summary_ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        cell = summary_ws.cell(row, start)
        cell.fill, cell.font, cell.alignment = styles["subheader"]
        if formula: cell.value = f"={formula}"
    
    weights = f"=SUM({'+'.join(f'I{r}*(C{r}/100)' for r in range(start_row, row))})"
    summary_ws[f"I{row}"] = weights
    summary_ws[f"I{row}"].number_format = "0.00%"
    summary_ws[f"D7"] = f"=I{row}"
    summary_ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)

    # Ranges Table
    ranges_header = row + 2
    summary_ws.merge_cells(f"B{ranges_header}:Q{ranges_header}")
    summary_ws[f"B{ranges_header}"] = "RANGES"
    summary_ws[f"B{ranges_header}"].fill, summary_ws[f"B{ranges_header}"].font, summary_ws[f"B{ranges_header}"].alignment = styles["header"]

    ranges_categories = ["CONCENTRATION", "pH", "TEMPERATURE", "PRESSURE", "FLOW RATE"]
    category_blocks = [(3, 5), (6, 8), (9, 11), (12, 14), (15, 17)]
    for cat, (start, end) in zip(ranges_categories, category_blocks):
        summary_ws.merge_cells(start_row=ranges_header+1, start_column=start, end_row=ranges_header+1, end_column=end)
        cell = summary_ws.cell(ranges_header+1, start)
        cell.value, cell.fill, cell.font, cell.alignment = cat, *styles["subheader"]

    for offset, sub in enumerate(["MIN", "MAX", "AVERAGE"]):
        for start, _ in category_blocks:
            cell = summary_ws.cell(ranges_header+2, start + offset)
            cell.value = sub
            cell.fill, cell.font, cell.alignment = light_fill, Font(color="000000", bold=True), Alignment(horizontal="center", vertical="center")

    r = ranges_header + 3
    for chem in chemicals:
        chem_df = df[df["CHEMICAL"] == chem]
        summary_ws.cell(r, 2, chem).fill = styles["light"][0]
        for i, col in enumerate(["CONCENTRATION", "pH LEVEL", "TEMPERATURE", "PRESSURE", "FLOW RATE"]):
            start = category_blocks[i][0]
            for j, val in enumerate([chem_df[col].min(), chem_df[col].max(), chem_df[col].mean()]):
                summary_ws.cell(r, start + j).value = val
                summary_ws.cell(r, start + j).alignment = Alignment(horizontal="center", vertical="center")
        r += 1

    # Column Widths and Borders
    summary_ws.column_dimensions["A"].width = 0.7
    summary_ws.column_dimensions["B"].width = 14
    for col in range(3, 18):
        summary_ws.column_dimensions[get_column_letter(col)].width = 9.4

    def apply_thick_border(ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                borders = {k: cell.border.__dict__[k] for k in ("left", "right", "top", "bottom")}
                if cell.row == rows[0][0].row: borders["top"] = thick
                if cell.row == rows[-1][0].row: borders["bottom"] = thick
                if cell.column == rows[0][0].column: borders["left"] = thick
                if cell.column == rows[0][-1].column: borders["right"] = thick
                cell.border = Border(**borders)

    for rng in ["B1:Q1", "B3:E7", "N3:Q7", "B9:Q9", f"C10:Q{row}", f"B{ranges_header}:Q{ranges_header+len(chemicals)+2}"]:
        apply_thick_border(summary_ws, rng)

    for start, end in category_blocks:
        apply_thick_border(summary_ws, f"{get_column_letter(start)}{ranges_header+1}:{get_column_letter(end)}{r-1}")

    # Ensure Summary is the active sheet
    wb.active = wb["Summary"]
    
    wb.save(save_path)
    print(f"Final formatted report saved at: {save_path}")


# -------------------- Main --------------------
def main():
    csv_path = select_file()
    if not csv_path:
        print("No file selected. Exiting.")
        return

    df = pd.read_csv(csv_path)
    user_info = get_user_info()
    save_path = get_save_path(user_info)
    if not save_path:
        print("No save location chosen. Exiting.")
        return

    df = standardize_csv_headers(df)
    ranges_df = load_ranges()
    df = check_compliance(df, ranges_df)
    format_excel(df, save_path, user_info)

if __name__ == "__main__":
    main()
